"""Build migration/acuity-import.xlsx — the master Acuity migration workbook.

Run: migration/.venv/bin/python migration/build_workbook.py
Output: migration/acuity-import.xlsx
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path(__file__).parent / "acuity-import.xlsx"

# ---------------------------------------------------------------------------
# Styling palette
# ---------------------------------------------------------------------------
BRAND_DARK = "1F2937"
BRAND_ACCENT = "8B2540"
HEADER_FILL = PatternFill("solid", fgColor=BRAND_DARK)
HEADER_FONT = Font(name="Inter", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Inter", size=18, bold=True, color=BRAND_DARK)
SUBTITLE_FONT = Font(name="Inter", size=11, italic=True, color="6B7280")
BODY_FONT = Font(name="Inter", size=10)
BODY_BOLD = Font(name="Inter", size=10, bold=True)
NOTE_FONT = Font(name="Inter", size=9, italic=True, color="9CA3AF")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(border_style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CATEGORY_COLORS = {
    "Semi-Permanent Make-up": "8B2540",
    "Saline Tattoo Removal": "475569",
    "Lashes & Brows": "D4747A",
    "Facials": "7B9B7E",
    "Waxing": "C89B47",
}

# ---------------------------------------------------------------------------
# DATA — appointment types (cleaned, normalised, in display order)
# ---------------------------------------------------------------------------
# Each row: (name, description, post_book_message, duration, pad_before,
#            pad_after, price, category, intake_form, image_note, iggy_notes)
APPOINTMENT_TYPES: list[tuple] = [
    (
        'Free Consultation & Patch Test',
        'Skin suitability assessment for semi-permanent brow procedures. We discuss your goals and expectations, perform a colour match, and carry out an allergy patch test.',
        '',
        15, 0, 5, 0,
        'Semi-Permanent Make-up',
        'SPMU Intake Form (required)',
        'Use SPMU consultation hero image',
        'Booksolo had duration 0 — set to 15min. Free, used as gateway to all SPMU services.',
    ),
    (
        'Beauty Spots',
        'A small artificial beauty mark applied with permanent cosmetics — face, shoulder, neck, or anywhere on the body. Permanent placement of a single dark spot.',
        'Avoid water and skincare on the treated area for 24 hours. Aftercare instructions will be provided.',
        60, 15, 15, 35,
        'Semi-Permanent Make-up',
        'SPMU Intake Form (required)',
        '',
        'Patch test required first — verify client has completed Free Consultation.',
    ),
    (
        'Freckles',
        'Semi-permanent freckle tattoos — soft, natural-looking pigment dots that fade over 1–3 years.',
        'Avoid water and skincare on the treated area for 24 hours. Aftercare instructions will be provided.',
        60, 15, 15, 70,
        'Semi-Permanent Make-up',
        'SPMU Intake Form (required)',
        '',
        "Pricing on Booksolo shown as 'from £70' — confirm whether this is a starting price or fixed.",
    ),
    (
        'Brow Microblading',
        'Semi-permanent technique for enhancing the appearance of the eyebrows, in which pigment is implanted into the skin in fine, short strokes resembling hair, using a hand tool with a blade formed of tiny needles. Shape and style is adapted to every client individually following their face features, skin and preferences.\n\nIn the price included :\n• Free consultation and a patch test\n• Initial procedure\n• Touch up in 5-8 weeks\n• After care kit\n\nNOTE: please book consultation and a patch test appointment first.',
        'Avoid getting brows wet for 7 days. Full aftercare guide will be emailed to you. Touch-up appointment is included — please book within 5–8 weeks.',
        120, 15, 15, 295,
        'Semi-Permanent Make-up',
        'SPMU Intake Form (required)',
        "Use Iggy's microblading portfolio shot",
        "Price INCLUDES 5–8wk touch-up — recommend creating a private 'SPMU Touch-Up (Included)' appt type for booking that follow-up.",
    ),
    (
        'Ombré Brows',
        'Semi- permanent make up treatment that uses either a digital machine to insert tiny dots of pigment into the skin. It creates a more shaded effect, rather than the crisp hair strokes that are often associated with microblading. Shape and style is adapted to every client individually following their face features, skin and preferences.\n\nIn the price included :\n• Free consultation and a patch test\n• Initial procedure\n• Touch up in 5-8 weeks\n• After care kit\n\nNOTE: please book consultation and a patch test appointment first.',
        'Avoid getting brows wet for 7 days. Full aftercare guide will be emailed to you. Touch-up appointment is included — please book within 5–8 weeks.',
        120, 15, 15, 295,
        'Semi-Permanent Make-up',
        'SPMU Intake Form (required)',
        'Use ombré-brows portfolio shot',
        "Price INCLUDES 5–8wk touch-up. OWNER'S SHEET SAID £15 — assumed typo, restored £295. CONFIRM before entering.",
    ),
    (
        'Combination Brows',
        'The Combination Brows is a treatment that combines hair stroke simulation with a mist of colour using shading techniques with digital machine. Combining microblading with ombre powder together creates a more defined looking brow than just hairstrokes alone. Shape and style is adapted to every client individually following their face features, skin and preferences.\n\nIn the price included :\n• Free consultation and a patch test\n• Initial procedure\n• Touch up in 5-8 weeks\n• After care kit\n\nNOTE: please book consultation and a patch test appointment first.',
        'Avoid getting brows wet for 7 days. Full aftercare guide will be emailed to you. Touch-up appointment is included — please book within 5–8 weeks.',
        120, 15, 15, 320,
        'Semi-Permanent Make-up',
        'SPMU Intake Form (required)',
        'Use combination-brows portfolio shot',
        'Price INCLUDES 5–8wk touch-up. Highest-margin SPMU service.',
    ),
    (
        'SPMU Brow Colour Boost — up to 6 months',
        'Touch-up for existing clients within 6 months of their original treatment. Refreshes colour saturation while pigment is still strong.',
        'Standard aftercare applies. Avoid getting brows wet for 7 days.',
        60, 15, 15, 90,
        'Semi-Permanent Make-up',
        'SPMU Intake Form (required)',
        '',
        'Client must have had original procedure with Iggy.',
    ),
    (
        'SPMU Brow Colour Boost — up to 9 months',
        'Touch-up for existing clients within 9 months of their original treatment. Light pigment refresh and shape correction.',
        'Standard aftercare applies. Avoid getting brows wet for 7 days.',
        60, 15, 15, 100,
        'Semi-Permanent Make-up',
        'SPMU Intake Form (required)',
        '',
        'Client must have had original procedure with Iggy.',
    ),
    (
        'SPMU Brow Colour Boost — up to 12 months',
        'Annual maintenance touch-up. Restores pigment depth and adjusts shape to suit your face today.',
        'Standard aftercare applies. Avoid getting brows wet for 7 days.',
        60, 15, 15, 120,
        'Semi-Permanent Make-up',
        'SPMU Intake Form (required)',
        '',
        'Client must have had original procedure with Iggy.',
    ),
    (
        'SPMU Brow Colour Boost — up to 18 months',
        'Maintenance touch-up for clients up to 18 months post-procedure. Includes more substantial re-pigmentation work.',
        'Standard aftercare applies. Avoid getting brows wet for 7 days.',
        60, 15, 15, 135,
        'Semi-Permanent Make-up',
        'SPMU Intake Form (required)',
        '',
        'Client must have had original procedure with Iggy.',
    ),
    (
        'SPMU Brow Correction or Colour Boost — over 18 months',
        'Full correction or re-pigmentation for clients whose original treatment was over 18 months ago. May include shape redesign.',
        'Standard aftercare applies. Avoid getting brows wet for 7 days.',
        60, 15, 15, 150,
        'Semi-Permanent Make-up',
        'SPMU Intake Form (required)',
        '',
        'Open to clients whose previous SPMU was with another artist — Iggy assesses suitability.',
    ),
    (
        'Saline Tattoo Removal — Single Session',
        'Lightens or removes existing semi-permanent makeup using a saline solution to draw pigment back out of the skin. Each client typically needs 3–6 sessions spaced 6–8 weeks apart.',
        'Avoid water on the treated area for 7 days. Full aftercare will be provided.',
        60, 15, 15, 45,
        'Saline Tattoo Removal',
        'SPMU Intake Form (required)',
        '',
        'Packages of 3 or 6 sessions sold separately on the Packages tab — Iggy applies the package code at checkout.',
    ),
    (
        'Eyebrow Shaping',
        'Wax, tweeze, and trim to define and tidy the brow shape.',
        'Avoid tanning, excessive sweating, steam room, sauna, swimming and makeup for 24 hours.',
        30, 5, 5, 10,
        'Lashes & Brows',
        '',
        '',
        '',
    ),
    (
        "Men's Eyebrow Shaping",
        'Wax, tweeze, and trim to tidy and neaten masculine brow shape without over-defining.',
        'Avoid tanning, excessive sweating, steam room, sauna, swimming and makeup for 24 hours.',
        30, 5, 5, 10,
        'Lashes & Brows',
        '',
        '',
        '',
    ),
    (
        'Eyebrow Tint',
        'Colour and depth added to brow hairs and skin using professional brow tint.',
        '',
        30, 5, 5, 10,
        'Lashes & Brows',
        'Tint Patch Test Confirmation',
        '',
        'Acuity should warn new clients to book a patch test first — handle via intake form.',
    ),
    (
        'Lash Tint',
        'Adds depth, shine, and definition to natural lashes — particularly useful for fair lash colour.',
        '',
        30, 5, 5, 15,
        'Lashes & Brows',
        'Tint Patch Test Confirmation',
        '',
        '',
    ),
    (
        'Eyebrow Tint & Wax',
        'Combination treatment — brow tint plus shaping wax in a single appointment.',
        'Avoid tanning, excessive sweating, steam room, sauna, swimming and makeup for 24 hours.',
        30, 5, 5, 20,
        'Lashes & Brows',
        'Tint Patch Test Confirmation',
        '',
        '',
    ),
    (
        'Lash & Eyebrow Tint',
        'Combined lash and brow tint to deepen colour on both areas in one visit.',
        '',
        30, 5, 5, 25,
        'Lashes & Brows',
        'Tint Patch Test Confirmation',
        '',
        '',
    ),
    (
        'Lash Tint, Eyebrow Tint & Wax',
        'Complete brow-and-lash refresh — lash tint, brow tint, and brow shaping in one appointment.',
        'Avoid tanning, excessive sweating, steam room, sauna, swimming and makeup for 24 hours.',
        30, 5, 5, 30,
        'Lashes & Brows',
        'Tint Patch Test Confirmation',
        '',
        '30min may be tight for this combo — verify with Iggy.',
    ),
    (
        'Lash Lift & Tint',
        'Lifts and curls natural lashes at the root for a wide-eyed effect, finished with a tint for added depth.',
        'Keep Lashes Dry: Avoid water, steam, and moisture around your lashes for at least 24-48 hours after the treatment to allow the lift and tint to set.\nAvoid Heat and Sweat: Stay away from saunas, hot showers, steam rooms, and exercise that causes sweating during this time.\nNo Rubbing or Touching: Do not rub, tug, or touch your lashes excessively to prevent disrupting the shape. No Eyelash Curlers: Do not use mechanical eyelash curlers, as this can damage the lifted lashes.\nCondition Lashes: After 48 hours, you can apply a lash conditioning serum. Brush Lashes: Use a clean spoolie to gently comb your lashes each morning to maintain their shape and separation.\nAvoid Excessive Heat: Keep lashes away from extreme heat, such as blow dryers, which could affect the curl.',
        45, 5, 10, 45,
        'Lashes & Brows',
        'Tint Patch Test Confirmation',
        '',
        '',
    ),
    (
        'Classic Lash Extensions — Full Set',
        'Individual lash extensions applied one-to-one to natural lashes for a fuller, longer look.',
        'Avoid Getting Lashes Wet: Keep your lashes completely dry for the first 24-48 hours to allow the adhesive to properly set.\nNo Steam or Heat: Avoid saunas, steam rooms, or hot showers during this time, as steam can weaken the adhesive.\nDo Not Touch or Rub: Refrain from touching, rubbing, or pulling on the lashes, as this can cause them to fall out or become misaligned.\nBe Gentle When Cleansing: Use a lash extension-safe cleanser to gently clean your eyes with a soft brush or cotton pad.\nAvoid Oil-Based Products: Stay away from oil-based cleansers, makeup removers, and moisturizers around the eye area, as they can break down the adhesive.\nBrush Your Lashes: Use a clean spoolie brush to gently comb through your lashes each morning to keep them tidy and in place.\nAvoid Excessive Heat: Keep away from blow dryers, ovens, or any high heat directly on your lashes.\nSchedule Refills: Lash extensions naturally shed with your natural lashes. Schedule refills every 2-3 weeks to maintain fullness and health of your lashes.',
        90, 10, 10, 60,
        'Lashes & Brows',
        '',
        "Use Iggy's lash portfolio shot",
        '',
    ),
    (
        'Classic Lash Infill — 2 Weeks',
        'Maintenance infill for classic lash extensions, booked within 2 weeks of last appointment. Outgrown extensions are removed and new ones applied.',
        'Avoid water and steam for 24 hours.',
        30, 10, 5, 35,
        'Lashes & Brows',
        '',
        '',
        'Acuity ideally restricts this to clients with a previous Classic full set — handled manually for now.',
    ),
    (
        'Classic Lash Infill — 3 Weeks',
        'Maintenance infill for classic lash extensions, booked within 3 weeks. More fill required than 2-week infill.',
        'Avoid water and steam for 24 hours.',
        45, 10, 5, 45,
        'Lashes & Brows',
        '',
        '',
        '',
    ),
    (
        'Hybrid Lash Extensions — New Set',
        'Mix of individual classic lashes and small fan-style volume lashes for a textured, mid-density look.',
        'Avoid Getting Lashes Wet: Keep your lashes completely dry for the first 24-48 hours to allow the adhesive to properly set.\nNo Steam or Heat: Avoid saunas, steam rooms, or hot showers during this time, as steam can weaken the adhesive.\nDo Not Touch or Rub: Refrain from touching, rubbing, or pulling on the lashes, as this can cause them to fall out or become misaligned.\nBe Gentle When Cleansing: Use a lash extension-safe cleanser to gently clean your eyes with a soft brush or cotton pad.\nAvoid Oil-Based Products: Stay away from oil-based cleansers, makeup removers, and moisturizers around the eye area, as they can break down the adhesive.\nBrush Your Lashes: Use a clean spoolie brush to gently comb through your lashes each morning to keep them tidy and in place.\nAvoid Excessive Heat: Keep away from blow dryers, ovens, or any high heat directly on your lashes.\nSchedule Refills: Lash extensions naturally shed with your natural lashes. Schedule refills every 2-3 weeks to maintain fullness and health of your lashes.',
        90, 10, 10, 65,
        'Lashes & Brows',
        '',
        '',
        '',
    ),
    (
        'Hybrid Lash Infill — 2 Weeks',
        'Maintenance infill for hybrid lash extensions, booked within 2 weeks of last appointment.',
        'Avoid water and steam for 24 hours.',
        30, 10, 5, 40,
        'Lashes & Brows',
        '',
        '',
        '',
    ),
    (
        'Hybrid Lash Infill — 3 Weeks',
        'Maintenance infill for hybrid lash extensions, booked within 3 weeks.',
        'Avoid water and steam for 24 hours.',
        45, 10, 5, 50,
        'Lashes & Brows',
        '',
        '',
        '',
    ),
    (
        'Volume Lash Extensions — New Set',
        'Fans of ultra-fine lash extensions applied to each natural lash for the densest, dramatic look.',
        'Avoid Getting Lashes Wet: Keep your lashes completely dry for the first 24-48 hours to allow the adhesive to properly set.\nNo Steam or Heat: Avoid saunas, steam rooms, or hot showers during this time, as steam can weaken the adhesive.\nDo Not Touch or Rub: Refrain from touching, rubbing, or pulling on the lashes, as this can cause them to fall out or become misaligned.\nBe Gentle When Cleansing: Use a lash extension-safe cleanser to gently clean your eyes with a soft brush or cotton pad.\nAvoid Oil-Based Products: Stay away from oil-based cleansers, makeup removers, and moisturizers around the eye area, as they can break down the adhesive.\nBrush Your Lashes: Use a clean spoolie brush to gently comb through your lashes each morning to keep them tidy and in place.\nAvoid Excessive Heat: Keep away from blow dryers, ovens, or any high heat directly on your lashes.\nSchedule Refills: Lash extensions naturally shed with your natural lashes. Schedule refills every 2-3 weeks to maintain fullness and health of your lashes.',
        90, 10, 10, 70,
        'Lashes & Brows',
        '',
        '',
        '',
    ),
    (
        'Volume Lash Infill — 2 Weeks',
        'Maintenance infill for volume lash extensions, booked within 2 weeks of last appointment.',
        'Avoid water and steam for 24 hours.',
        30, 10, 5, 40,
        'Lashes & Brows',
        '',
        '',
        '',
    ),
    (
        'Volume Lash Infill — 3 Weeks',
        'Maintenance infill for volume lash extensions, booked within 3 weeks.',
        'Avoid water and steam for 24 hours.',
        45, 10, 5, 50,
        'Lashes & Brows',
        '',
        '',
        '',
    ),
    (
        'Lash Extension Removal',
        'Professional removal of all lash extensions without damage to the natural lashes.',
        '',
        30, 5, 5, 20,
        'Lashes & Brows',
        '',
        '',
        '',
    ),
    (
        "Men's Facial",
        'The Men’s Facial is a revitalizing treatment tailored specifically for men’s skin, targeting issues like clogged pores, rough texture, and irritation from shaving. This facial includes a deep cleanse, exfoliation to remove dead skin cells, and a soothing mask to reduce redness and inflammation. It’s designed to hydrate, smooth, and refresh the skin, addressing common concerns such as ingrown hairs and environmental stress. Ideal for all skin types, this treatment leaves the face feeling clean, energized, and healthier-looking. Perfect for men who want a straightforward yet effective skincare experience.',
        'Arrive with a clean-shaven face if possible, or with day-old growth — avoid shaving immediately before.',
        65, 10, 10, 65,
        'Facials',
        '',
        '',
        'Booksolo lists 65min/£65 — verify duration with Iggy (round 60min or 75min more typical).',
    ),
    (
        'Deep Cleansing Facial',
        'A deep cleansing facial is a thorough skin treatment designed to remove impurities, toxins, and dead skin cells from the face. It typically involves cleansing, exfoliation, steam, extractions, and a mask suited to each individual skin needs, leaving the skin refreshed, clear, and rejuvenated. Ideal for congested skin, this facial helps to unclog pores, reduce breakouts, and improve overall skin texture and tone. It can be the starter facial to set a treatment plan and prepare your skin for more advanced facials.',
        'SPF50 for all clients\nNo swimming for 48h\n-No excessive exercise, heat treatments and makeup for 48 hours\nNo other facial treatments for at least one week\nStay well hydrated',
        60, 10, 10, 60,
        'Facials',
        '',
        '',
        '',
    ),
    (
        'Dermaplaning',
        'Dermaplaning is a manual exfoliation procedure that gently abrades the dead layers of the skin using a surgical blade. The procedure aims to increase the absorption of skincare products and diminish fine lines and wrinkles. The treatment creates a healthy glow and a radiant complexion.\nDermaplane aims to make your skins surface smooth, youthful and radiant. It is also used to remove "peach fuzz" the short soft vellus nonterminal hair from your face.\n. Perfect for clients with congested, dry and dull skin.\n- For those requiring a deep exfoliation.\n- Not suitable for those with active breakouts',
        'SPF50 for all clients\nNo swimming for 7 days\n-No excessive exercise, heat treatments and makeup for 48 hours\nNo exfoliation for 7 days\nNo other facial treatments for at least 2 weeks\nAvoid products such as Retinol for 7 days before/ after\nStay well hydrated',
        50, 10, 10, 50,
        'Facials',
        '',
        '',
        '',
    ),
    (
        'Microneedling',
        'Microneedling is a treatment used on the face& bodyspecifically targeting acne scarring, pore size and fine lines & wrinkles. The act of Microneedling creates multiple superficial puncture wounds, releasing growth factors\nthat stimulate the fibroblasts creating new collagen and elastin.\nThe aim of the treatment is to traumatise the skin without collateral damage to the epidermis and healthy tissue.\nThe recovery is rapid with very little downtime, residual pain or discomfort. Microneedling increases the availability of cell nutrients and helps to strengthen the collagen in vessel walls and connective tissue.\nIt increases the density of the epidermis lessening the appearance of broken capillaries.\n\nBENFFITS\n• Minimises fine lines and wrinkles\n• Improves acne scarring\n• Minimises pore size\n• Improves texture and tone\n• Rejuvenates the skin\n• Encourages the production of collagen\n\nNot suitable for clients with active breakouts.',
        'SPF50 for all clients\nNo swimming for 7 days\n-No excessive exercise, heat treatments and makeup for 48 hours\nNo exfoliation for 7 days\nNo other facial treatments for at least 2 weeks\nAvoid products such as Retinol for 7 days before/ after\nStay well hydrated',
        80, 15, 15, 80,
        'Facials',
        'Skin Treatment Consent Form',
        '',
        "Pairs with the 'Microneedling: Skin Boosters' add-on (see Add-ons tab).",
    ),
    (
        'Million Dollar Facial',
        'Million Dollar Facial is a luxurious facial where science meets indulgence. Working the skin from the outside in, it has been designed to deeply exfoliate dead skin cells, remove non-terminal hair, flush toxins, increase cell turnover and stimulate collagen production.\nOur Million-dollar facial treatment is a 10-step unique protocol which combines dermaplaning, skin needling and lymphatic drainage massage to leave clients looking and feeling a million dollars.\n- This facial will help to target pigmentation, fine lines, wrinkles, dull skin and scarring.\n- Combats and reduces signs of aging revealing a smooth and more even toned complexion.\n- Not suitable for clients with active breakouts.',
        'Avoid sun exposure for 48 hours. Full at-home regimen will be discussed.',
        100, 15, 15, 100,
        'Facials',
        'Skin Treatment Consent Form',
        'Use Million Dollar branded imagery — Iggy is licensed',
        'Iggy holds a Million Dollar Facial license — verify branding compliance.',
    ),
    (
        'Immaculate Peel',
        'A chemical peel enhances and smoothens the texture of the skin. It is an effective treatment for facial blemishes, wrinkles and uneven pigmentation, problematic skin and pustules. They exfoliate the outer layers of dead skin, revealing a new skin layer with improved tone, texture and colour.\n\nIt is designed to give near immediate results without the downtime often associated with chemical peels. Immaculate Peels have been developed to support improvements to problematic, inflamed, irritated oily or spot prone skin types to help calm, heal and restore a more-healthy calm appearance.\n\nCONCERNS\n• Active breakouts, blackheads, comedones, rough texture and enlarged pores\n• Photodamaged skin, acne scarring, pigmentation, dull skin, fine lines\n• Aging skin, loss of elasticity, pigmentation, age spots, dry skin, oxative stress ( lifestyle& pollution ), fine lines & wrinkles\n\n\nBENEFITS OF CHEMICAL PEELS\n• One of the most popular non-invasive treatments available.\n• No downtime, no peeling.\n• Speedy treatment.\n• Powerful antioxidant & antimicrobial properties.\n\nCan help:\n• Fine lines,\n• Uneven skin tone and texture,\n• Pigmentation and photodamage.',
        'SPF50 for all clients\nNo swimming for 7 days\n-No excessive exercise, heat treatments and makeup for 48 hours\nNo exfoliation for 7 days\nNo other facial treatments for at least 2 weeks\nAvoid products such as Retinol for 7 days before/ after\nStay well hydrated',
        55, 10, 10, 55,
        'Facials',
        'Skin Treatment Consent Form',
        '',
        '',
    ),
    (
        'PregnaCare Facial',
        'The Pampering Cleansing & Hydrating Facial for Pregnant Women is a gentle, nourishing treatment designed to address the unique skincare needs of expectant mothers. This facial uses pregnancy-safe products to deeply cleanse, hydrate, and soothe sensitive skin without any harsh chemicals. It includes a gentle cleanse, light exfoliation, pore cleanse, calming mask, and a hydrating serum to restore moisture, leaving skin soft, glowing, and refreshed. Perfect for relieving dryness, sensitivity, and hormonal skin changes, this facial is a safe, relaxing escape for moms-to-be looking for a radiant, healthy complexion.',
        "Please let us know which trimester you're in when booking.",
        55, 10, 10, 55,
        'Facials',
        'Pregnancy Skin Intake',
        '',
        'Consider a pregnancy-specific intake form to confirm trimester and any complications.',
    ),
    (
        'Million Dollar Miracle Mask',
        'Million Dollar Miracle Mask is a ground breaking facial mask packed with clinically proven ingredients. Leaving skin tighter, firmer and brighter.\n\nThe formula stimulates plasma into the deeper levels of the skin in order to enhance cellular renewal. Using the latest in peptide technology, this Miracle Mask has the unique ability to create micro circulation, promote collagen and detoxify skin. By helping to promote keratinocytes, collagen and elastin, the skin can remove impurities and provide a youthful appearance\n\nTreatment can be combined with Dermaplaning.\n\n- Suitable for all skin types\n- Perfect to have 72 hours before a special occasion\n- Leaves skin tighter, firmer and brighter',
        'SPF50 for all clients\nNo swimming for 48h\n-No excessive exercise, heat treatments and makeup for 48 hours\nNo other facial treatments for at least one week\nStay well hydrated',
        55, 10, 10, 55,
        'Facials',
        '',
        '',
        '',
    ),
    (
        'Back Cleansing Facial',
        'A back cleansing facial is a targeted treatment that focuses on deep cleaning the skin on the back, an area prone to breakouts, clogged pores, and uneven texture. This facial typically includes cleansing, exfoliation, steam, and extractions to remove impurities and excess oil, followed by a soothing mask to hydrate and calm the skin. It helps to clear acne, reduce blemishes, and leave the back smooth and refreshed. Ideal for those with back acne or anyone looking to improve the appearance of their skin in hard-to-reach areas.',
        'SPF50 for all clients\nNo swimming for 48h\n-No excessive exercise, heat treatments and makeup for 48 hours\nNo other facial treatments for at least one week\nStay well hydrated',
        60, 10, 10, 60,
        'Facials',
        '',
        '',
        '',
    ),
    (
        'Eyebrow Wax',
        'Shaping using wax, finished with tweezing and trimming where needed.',
        '',
        30, 5, 5, 10,
        'Waxing',
        '',
        '',
        'Booksolo lists 30min — many salons book 15min for this. Confirm with Iggy.',
    ),
    (
        "Men's Brow Wax",
        "Tidy and shape men's brows using wax, tweezers, and scissors for a clean, masculine finish.",
        '',
        30, 5, 5, 10,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Brow, Lip & Chin Wax',
        'Eyebrow wax with upper-lip wax and chin wax in one appointment.',
        '',
        30, 5, 5, 20,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Brow & Lip Wax',
        'Eyebrow wax combined with upper-lip wax.',
        '',
        30, 5, 5, 15,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Chin Wax',
        'Removal of any hair from the chin area using wax.',
        '',
        30, 5, 5, 6,
        'Waxing',
        '',
        '',
        'Duration likely 15min in reality — confirm.',
    ),
    (
        'Lip Wax',
        'Removal of upper-lip hair using wax.',
        '',
        30, 5, 5, 5,
        'Waxing',
        '',
        '',
        'Duration likely 15min in reality — confirm.',
    ),
    (
        'Nose Wax',
        'Removal of nose hair using hot wax.',
        '',
        30, 5, 5, 6,
        'Waxing',
        '',
        '',
        'Duration likely 15min in reality — confirm.',
    ),
    (
        'Facial Sides Wax',
        'Hair removal from the sides of the face using wax.',
        '',
        30, 5, 5, 8,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Underarm Wax',
        'Hair removal under the arms.',
        '',
        30, 5, 5, 10,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Half Arm Wax',
        'Hair removal from the forearm (elbow down) or upper arm.',
        '',
        30, 5, 5, 15,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Full or 3/4 Arm Wax',
        'Hair removal from the full or three-quarter length of the arm.',
        '',
        30, 5, 5, 20,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Half Leg Wax — Upper',
        'Hair removal from the thigh using strip wax.',
        '',
        30, 5, 5, 20,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Half Leg Wax — Lower',
        'Hair removal from the knee down using strip wax.',
        '',
        30, 5, 5, 15,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Full Leg Wax',
        'Complete leg hair removal — thigh to ankle — using strip wax.',
        '',
        30, 10, 10, 30,
        'Waxing',
        '',
        '',
        '30min may be tight for a full leg — verify with Iggy (45–60min typical).',
    ),
    (
        'Bikini Line Wax',
        'Hair removal from outside the knicker line.',
        '',
        30, 5, 5, 20,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Extended Bikini Wax',
        'Hair removal further beyond the standard knicker line.',
        '',
        30, 5, 5, 30,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Brazilian Bikini Wax',
        'Removal of most bikini hair, leaving a small landing strip.',
        '',
        30, 5, 10, 35,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Hollywood Bikini Wax',
        'Full removal of all bikini hair.',
        '',
        30, 5, 10, 35,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Stomach Wax',
        'Removal of stomach hair using strip wax.',
        '',
        30, 5, 5, 15,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Back Wax',
        'Removal of back hair using strip wax.',
        '',
        30, 5, 5, 20,
        'Waxing',
        '',
        '',
        '',
    ),
    (
        'Chest Wax',
        'Removal of chest hair using strip wax.',
        '',
        30, 5, 5, 15,
        'Waxing',
        '',
        '',
        '',
    ),
]

# ---------------------------------------------------------------------------
# Add-ons
# ---------------------------------------------------------------------------
ADDONS = [
    (
        "Premium Facial Add-on",
        "Additional premium treatment layered onto any facial — Iggy selects from collagen masks, hydrating boosters, or eye/lip treatments based on your skin needs.",
        30,    # extra minutes
        20,    # extra price
        "Public",
        ["Deep Cleansing Facial", "Dermaplaning", "Immaculate Peel", "Million Dollar Facial",
         "Million Dollar Miracle Mask", "Back Cleansing Facial", "Men's Facial", "PregnaCare Facial"],
        "Generic up-sell. Consider splitting into named add-ons (Collagen Mask £20, Hydrating Booster £20, etc.) for clarity.",
    ),
    (
        "Microneedling: Skin Boosters",
        "Enhance your microneedling treatment with Exosomes, Polynucleotides, or Collagen for accelerated rejuvenation.",
        30,
        40,
        "Public",
        ["Microneedling"],
        "Iggy may want this split into 3 separate add-ons (one per booster type) so clients pick which one.",
    ),
]

# ---------------------------------------------------------------------------
# Packages
# ---------------------------------------------------------------------------
PACKAGES = [
    (
        "Saline Tattoo Removal — 3 Session Package",
        "Three saline tattoo-removal sessions sold together at a discounted rate. Sessions spaced 6–8 weeks apart for best results.",
        120,                                # price £
        3,                                  # number of credits / sessions
        ["Saline Tattoo Removal — Single Session"],
        12,                                 # expiry months
        "£15 saving vs booking 3 single sessions (£135). Client receives a code to redeem each session.",
    ),
    (
        "Saline Tattoo Removal — 6 Session Package",
        "Six saline tattoo-removal sessions sold together at a discounted rate. Recommended for clients requiring full removal of denser pigment.",
        230,
        6,
        ["Saline Tattoo Removal — Single Session"],
        18,
        "£40 saving vs booking 6 single sessions (£270). Longer expiry to accommodate the 8-week spacing.",
    ),
]

# ---------------------------------------------------------------------------
# Categories master list
# ---------------------------------------------------------------------------
CATEGORIES = [
    ("Semi-Permanent Make-up",
     "Long-wear pigmentation services — brows, freckles, beauty spots.",
     "Burgundy", CATEGORY_COLORS["Semi-Permanent Make-up"], 11),
    ("Saline Tattoo Removal",
     "Pigment removal and correction service. Packaged in 3 or 6 sessions.",
     "Slate", CATEGORY_COLORS["Saline Tattoo Removal"], 1),
    ("Lashes & Brows",
     "Tinting, shaping, lifts, and lash extensions.",
     "Rose", CATEGORY_COLORS["Lashes & Brows"], 18),
    ("Facials",
     "Skin treatments including peels, dermaplaning, microneedling, and signature facials.",
     "Sage", CATEGORY_COLORS["Facials"], 9),
    ("Waxing",
     "Face and body waxing services.",
     "Amber", CATEGORY_COLORS["Waxing"], 21),
]

# ---------------------------------------------------------------------------
# Pre-flight checklist (things to do BEFORE adding services)
# ---------------------------------------------------------------------------
PREFLIGHT = [
    ("Account Setup", "Sign up for Acuity at acuityscheduling.com — Standard plan or higher is required for Packages.", "https://acuityscheduling.com"),
    ("Account Setup", "Set business name, address (82 O'Leary Square, London E1 3AS), phone, and contact email.", ""),
    ("Account Setup", "Set timezone to Europe/London (UK).", ""),
    ("Account Setup", "Set currency to GBP. This is one-time and CANNOT be changed without contacting Acuity support.", ""),
    ("Branding", "Upload FaceFrame logo and set brand colours (use the hex codes from the Categories tab).", ""),
    ("Branding", "Choose scheduling-page URL slug, e.g. faceframebeauty.acuityscheduling.com.", ""),
    ("Payments", "Confirm Stripe is connected (Iggy already has Stripe). Needed for card-on-file enforcement and Packages.", "https://help.acuityscheduling.com/hc/articles/16676914206477"),
    ("Payments", "Turn ON 'Require credit card to book' GLOBALLY — applies to every appointment type. No charge at booking; card is captured for cancellation enforcement only.", ""),
    ("Payments", "Leave auto-charge for missed appointments OFF — Iggy charges manually per case via the appointment record. Reason: she wants discretion to waive the fee when a client is genuinely ill or has an emergency.", ""),
    ("Payments", "Do NOT enable booking deposits — Iggy's policy is no upfront charge. Card is only charged if the cancellation rule is broken AND Iggy decides to charge.", ""),
    ("Payments", "Cancellation policy text to use on the booking page footer + confirmation email: 'A valid card is required to book. Cancellations or reschedules within 48 hours of your appointment, and no-shows, may be charged the full service price. Exceptions are made at the artist's discretion for genuine illness or emergencies — please get in touch as soon as you know.'", ""),
    ("Calendars", "Create one Calendar named 'Iggy' (the studio's only practitioner).", ""),
    ("Availability", "Set studio business hours.", ""),
    ("Availability", "Add UK bank holidays as blocked days for the year.", ""),
    ("Intake Forms", "Create 'SPMU Intake Form' — medical history, allergies, medications, prior procedures, contraindications.", "https://help.acuityscheduling.com/hc/articles/16676931038093"),
    ("Intake Forms", "Create 'Skin Treatment Consent Form' — for microneedling, peels, dermaplaning, Million Dollar Facial.", ""),
    ("Intake Forms", "Create 'Tint Patch Test Confirmation' — yes/no checkbox + date of patch test.", ""),
    ("Intake Forms", "Create 'Pregnancy Skin Intake' — trimester, complications, GP approval if needed.", ""),
    ("Notifications", "Customise booking confirmation email template with FaceFrame branding.", ""),
    ("Notifications", "Customise reminder email/SMS (24-hour reminder recommended).", ""),
    ("Notifications", "Customise post-appointment thank-you email — include link to leave a review.", ""),
    ("Compliance", "Confirm UK VAT registration status — if registered, enable VAT on all services.", ""),
    ("Compliance", "Add GDPR-compliant privacy notice to intake forms.", ""),
    ("Then & only then", "Move to the AppointmentTypes tab and begin adding services.", ""),
]


# ---------------------------------------------------------------------------
# Build the workbook
# ---------------------------------------------------------------------------
def style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER
    ws.row_dimensions[row].height = 32


def set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_title(ws, title: str, subtitle: str = "") -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT
        ws.row_dimensions[2].height = 20


def build_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    write_title(ws, "FaceFrame Beauty — Acuity Migration Workbook",
                "Generated to migrate services from Booksolo into Acuity Scheduling. Work through the tabs in order.")
    set_col_widths(ws, [4, 26, 80])

    rows = [
        ("", "", ""),
        ("", "How to use this workbook", ""),
        ("1.", "Read this tab", "Start here — understand the structure and quirks before clicking anything in Acuity."),
        ("2.", "Pre-flight tab", "Complete every checklist item BEFORE you start adding services. Account setup, payments, intake forms, etc."),
        ("3.", "Categories tab", "Reference list of the 5 service categories with their canonical names and brand hex colours. Categories are created inline as you add the first service in each one."),
        ("4.", "AppointmentTypes tab", "The main tab — 60 services. Open Acuity → Appointment Types → 'New Type of Service' and copy each row across. Columns are in the same order as the Acuity form."),
        ("5.", "Add-ons tab", "After all appointment types are in, go to Appointment Types → Add-ons → 'Create add-on'."),
        ("6.", "Packages tab", "Last — Sell → Packages, Gift Certificates & Subscriptions → 'New Package'."),
        ("7.", "Source-Booksolo tab", "Raw audit copy of the original Booksolo data. Don't edit — kept for traceability."),
        ("", "", ""),
        ("", "Form field map (Acuity 'New Type of Service' → workbook column)", ""),
        ("•", "Name", "Workbook column: Name"),
        ("•", "Description", "Workbook column: Description"),
        ("•", "Show a message after scheduling", "Workbook column: Post-booking Message"),
        ("•", "Duration", "Workbook column: Duration (min)"),
        ("•", "Padding", "Workbook columns: Pad Before / Pad After (min). Tick 'Block off extra time before or after the appointment'."),
        ("•", "Price", "Workbook column: Price (£)"),
        ("•", "Category", "Workbook column: Category. First time in each category, click 'Create a new Category' and use the exact name from the Categories tab."),
        ("•", "Color", "Workbook column: Colour (hex). Acuity's picker doesn't accept hex directly — match the closest swatch."),
        ("•", "Access", "Workbook column: Access. Default Public for all 60 services."),
        ("•", "Intake form", "Workbook column: Intake Form. Create the listed forms during pre-flight, then tick them here."),
        ("•", "Image", "Workbook column: Image Note. Suggested image to upload."),
        ("", "", ""),
        ("", "Key decisions baked in (review before importing)", ""),
        ("•", "Free Consultation duration set to 15min", "Booksolo had it at 0 — Acuity requires a duration."),
        ("•", "Saline 3 & 6 session bundles → Packages", "These were single 180min / 360min appointments on Booksolo — almost certainly a misuse. Re-modelled as Packages of 3/6 credits redeemable against the single saline session."),
        ("•", "Two 'Add on' items moved to Add-ons tab", "Booksolo treated them as appointment types — Acuity's Add-ons mechanic is the right fit."),
        ("•", "Padding added per service", "5min for waxing/tints, 10min for facials, 15min for SPMU and microneedling (preparation + cleanup)."),
        ("•", "Touch-up included in SPMU pricing", "Microblading/Ombré/Combination all include a 5–8 week touch-up. Recommend creating a private 'SPMU Touch-Up (Included)' type for booking those follow-ups."),
        ("•", "Suspicious durations flagged", "Several waxing services were 30min on Booksolo (lip, chin, nose) — probably should be 15min. Flagged in the 'Notes for Iggy' column."),
        ("", "", ""),
        ("", "Why no API auto-import?", ""),
        ("", "Note", "Acuity's public API is read-only for appointment types. You cannot POST new services. The only paths are: (a) manual dashboard entry, or (b) browser automation. This workbook supports path (a) — minimising the typing time."),
    ]
    for i, (a, b, c) in enumerate(rows, start=3):
        ws.cell(row=i, column=1).value = a
        ws.cell(row=i, column=2).value = b
        ws.cell(row=i, column=2).font = BODY_BOLD if b and not c else BODY_FONT
        ws.cell(row=i, column=3).value = c
        ws.cell(row=i, column=3).alignment = WRAP
        ws.cell(row=i, column=3).font = BODY_FONT
    ws.sheet_view.showGridLines = False


def build_preflight(wb: Workbook) -> None:
    ws = wb.create_sheet("Pre-flight")
    write_title(ws, "Pre-flight Checklist",
                "Complete every item BEFORE you start adding services on the AppointmentTypes tab.")
    headers = ["✓", "Section", "Action", "Link / Note"]
    header_row = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, header_row, len(headers))
    set_col_widths(ws, [5, 18, 70, 50])
    for i, (section, action, link) in enumerate(PREFLIGHT, start=header_row + 1):
        ws.cell(row=i, column=1).value = "☐"
        ws.cell(row=i, column=1).alignment = CENTER
        ws.cell(row=i, column=2).value = section
        ws.cell(row=i, column=2).font = BODY_BOLD
        ws.cell(row=i, column=3).value = action
        ws.cell(row=i, column=3).alignment = WRAP
        ws.cell(row=i, column=3).font = BODY_FONT
        ws.cell(row=i, column=4).value = link
        ws.cell(row=i, column=4).font = NOTE_FONT
        ws.cell(row=i, column=4).alignment = WRAP
        ws.row_dimensions[i].height = 28
        if section == "Then & only then":
            for c in range(1, 5):
                ws.cell(row=i, column=c).fill = PatternFill("solid", fgColor="FEF3C7")
                ws.cell(row=i, column=c).font = Font(name="Inter", size=10, bold=True, color="92400E")
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


def build_categories(wb: Workbook) -> None:
    ws = wb.create_sheet("Categories")
    write_title(ws, "Categories",
                "Canonical category names and brand colours. Created inline when you add the first service in each category.")
    headers = ["Category Name", "Description", "Brand Colour (label)", "Hex", "# Services"]
    header_row = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, header_row, len(headers))
    set_col_widths(ws, [30, 60, 22, 14, 12])
    for i, (name, desc, label, hex_, count) in enumerate(CATEGORIES, start=header_row + 1):
        ws.cell(row=i, column=1).value = name
        ws.cell(row=i, column=1).font = BODY_BOLD
        ws.cell(row=i, column=2).value = desc
        ws.cell(row=i, column=2).alignment = WRAP
        ws.cell(row=i, column=2).font = BODY_FONT
        ws.cell(row=i, column=3).value = label
        ws.cell(row=i, column=3).font = BODY_FONT
        ws.cell(row=i, column=4).value = f"#{hex_}"
        ws.cell(row=i, column=4).fill = PatternFill("solid", fgColor=hex_)
        ws.cell(row=i, column=4).font = Font(name="Inter", size=10, bold=True, color="FFFFFF")
        ws.cell(row=i, column=4).alignment = CENTER
        ws.cell(row=i, column=5).value = count
        ws.cell(row=i, column=5).alignment = CENTER
        ws.row_dimensions[i].height = 28
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


def build_appointment_types(wb: Workbook) -> None:
    ws = wb.create_sheet("AppointmentTypes")
    write_title(ws, "Appointment Types (60 services)",
                "Open Acuity → Appointment Types → 'New Type of Service'. Columns match the form order.")
    headers = [
        "#", "✓", "Name", "Description", "Post-booking Message",
        "Duration (min)", "Pad Before (min)", "Pad After (min)", "Price (£)",
        "Category", "Colour (hex)", "Access", "Intake Form to Attach",
        "Image Note", "Notes for Iggy",
    ]
    header_row = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, header_row, len(headers))
    set_col_widths(ws, [4, 5, 38, 60, 50, 10, 10, 10, 10, 24, 14, 10, 26, 28, 50])
    for idx, row in enumerate(APPOINTMENT_TYPES, start=1):
        (name, desc, post_msg, dur, pad_b, pad_a, price, cat,
         intake, img, notes) = row
        r = header_row + idx
        cells = [
            idx, "☐", name, desc, post_msg, dur, pad_b, pad_a, price,
            cat, f"#{CATEGORY_COLORS[cat]}", "Public", intake, img, notes,
        ]
        for c, val in enumerate(cells, start=1):
            cell = ws.cell(row=r, column=c)
            cell.value = val
            cell.alignment = WRAP if c in (3, 4, 5, 13, 14, 15) else (CENTER if c in (1, 2, 6, 7, 8, 9, 11, 12) else Alignment(vertical="top"))
            cell.font = BODY_FONT
            cell.border = BORDER
        ws.cell(row=r, column=3).font = BODY_BOLD
        colour_cell = ws.cell(row=r, column=11)
        colour_cell.fill = PatternFill("solid", fgColor=CATEGORY_COLORS[cat])
        colour_cell.font = Font(name="Inter", size=9, bold=True, color="FFFFFF")
        ws.cell(row=r, column=15).font = NOTE_FONT
        ws.row_dimensions[r].height = 60

    cat_validation = DataValidation(
        type="list",
        formula1='"' + ",".join(CATEGORY_COLORS.keys()) + '"',
        allow_blank=False,
    )
    cat_validation.add(f"J{header_row + 1}:J{header_row + len(APPOINTMENT_TYPES)}")
    ws.add_data_validation(cat_validation)
    access_validation = DataValidation(
        type="list", formula1='"Public,Private"', allow_blank=False,
    )
    access_validation.add(f"L{header_row + 1}:L{header_row + len(APPOINTMENT_TYPES)}")
    ws.add_data_validation(access_validation)
    ws.freeze_panes = f"D{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:O{header_row + len(APPOINTMENT_TYPES)}"
    ws.sheet_view.showGridLines = False


def build_addons(wb: Workbook) -> None:
    ws = wb.create_sheet("Add-ons")
    write_title(ws, "Add-ons (2 items)",
                "Acuity → Appointment Types → Add-ons → 'Create add-on'. Tick the listed appointment types to attach each add-on.")
    headers = [
        "#", "✓", "Name", "Description", "Extra Time (min)", "Extra Price (£)",
        "Access", "Attach to Appointment Types", "Notes for Iggy",
    ]
    header_row = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, header_row, len(headers))
    set_col_widths(ws, [4, 5, 36, 60, 14, 14, 12, 60, 60])
    for idx, (name, desc, t, p, access, attached, notes) in enumerate(ADDONS, start=1):
        r = header_row + idx
        cells = [idx, "☐", name, desc, t, p, access, ", ".join(attached), notes]
        for c, val in enumerate(cells, start=1):
            cell = ws.cell(row=r, column=c)
            cell.value = val
            cell.alignment = WRAP if c in (3, 4, 8, 9) else (CENTER if c in (1, 2, 5, 6, 7) else Alignment(vertical="top"))
            cell.font = BODY_FONT
            cell.border = BORDER
        ws.cell(row=r, column=3).font = BODY_BOLD
        ws.cell(row=r, column=9).font = NOTE_FONT
        ws.row_dimensions[r].height = 80
    ws.freeze_panes = f"D{header_row + 1}"
    ws.sheet_view.showGridLines = False


def build_packages(wb: Workbook) -> None:
    ws = wb.create_sheet("Packages")
    write_title(ws, "Packages (2 bundles)",
                "Acuity → Sell → Packages, Gift Certificates & Subscriptions → 'New Package'. Requires Standard plan or higher and a connected payment processor.")
    headers = [
        "#", "✓", "Name", "Description", "Price (£)", "Sessions Included",
        "Redeemable Against", "Expires After (months)", "Notes for Iggy",
    ]
    header_row = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, header_row, len(headers))
    set_col_widths(ws, [4, 5, 38, 60, 12, 18, 36, 16, 60])
    for idx, (name, desc, price, sessions, against, expires, notes) in enumerate(PACKAGES, start=1):
        r = header_row + idx
        cells = [idx, "☐", name, desc, price, sessions, ", ".join(against), expires, notes]
        for c, val in enumerate(cells, start=1):
            cell = ws.cell(row=r, column=c)
            cell.value = val
            cell.alignment = WRAP if c in (3, 4, 7, 9) else (CENTER if c in (1, 2, 5, 6, 8) else Alignment(vertical="top"))
            cell.font = BODY_FONT
            cell.border = BORDER
        ws.cell(row=r, column=3).font = BODY_BOLD
        ws.cell(row=r, column=9).font = NOTE_FONT
        ws.row_dimensions[r].height = 80
    ws.freeze_panes = f"D{header_row + 1}"
    ws.sheet_view.showGridLines = False


def build_source(wb: Workbook) -> None:
    """Raw Booksolo data, preserved for audit. Do not edit."""
    ws = wb.create_sheet("Source-Booksolo")
    write_title(ws, "Source data — Booksolo (raw)",
                "Verbatim from https://booksolo.co/faceframebeauty/book/treatments — preserved for audit. Do not edit.")
    headers = ["#", "Name", "Category", "Duration (min)", "Price (£)", "Description"]
    header_row = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i).value = h
    style_header(ws, header_row, len(headers))
    set_col_widths(ws, [4, 50, 28, 14, 12, 80])

    from booksolo_raw import RAW
    for idx, item in enumerate(RAW, start=1):
        r = header_row + idx
        ws.cell(row=r, column=1).value = idx
        ws.cell(row=r, column=2).value = item["name"]
        ws.cell(row=r, column=3).value = item["category"]
        ws.cell(row=r, column=4).value = item["duration"]
        ws.cell(row=r, column=5).value = item["price"]
        ws.cell(row=r, column=6).value = item["description"]
        for c in range(1, 7):
            ws.cell(row=r, column=c).font = BODY_FONT
            ws.cell(row=r, column=c).alignment = WRAP if c in (2, 6) else CENTER
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 36
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:F{header_row + len(RAW)}"
    ws.sheet_view.showGridLines = False


def main() -> None:
    wb = Workbook()
    build_readme(wb)
    build_preflight(wb)
    build_categories(wb)
    build_appointment_types(wb)
    build_addons(wb)
    build_packages(wb)
    build_source(wb)
    wb.save(OUT)
    print(f"Wrote {OUT} — {len(APPOINTMENT_TYPES)} appointment types, "
          f"{len(ADDONS)} add-ons, {len(PACKAGES)} packages.")


if __name__ == "__main__":
    main()
